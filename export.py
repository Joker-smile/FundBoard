"""
导出功能模块

提供基金数据的 Excel 和 CSV 导出功能。
- Excel 导出使用 openpyxl，支持格式化（加粗标题、列宽自适应、涨跌着色）
- CSV 导出使用 csv 模块，UTF-8 BOM 编码以兼容 Excel 打开
"""

import csv
import os
from typing import Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# 列标题映射
COLUMN_HEADERS = [
    ("code", "基金代码"),
    ("name", "基金名称"),
    ("purchase_limit", "交易状态"),
    ("index_type", "跟踪指数"),
    ("nav", "单位净值"),
    ("nav_date", "净值日期"),
    ("acc_nav", "累计净值"),
    ("daily_change", "日增长值"),
    ("daily_change_pct", "日增长率(%)"),
    ("data_source", "数据来源"),
]


def export_to_excel(funds: List[Dict], filepath: str) -> bool:
    """
    导出基金数据到 Excel 文件。

    Args:
        funds: 基金数据列表
        filepath: 保存文件路径

    Returns:
        True 表示成功，False 表示失败
    """
    if not HAS_OPENPYXL:
        raise ImportError("需要安装 openpyxl 库: pip install openpyxl")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "基金数据"

        # === 写入标题行 ===
        header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, (_, header_text) in enumerate(COLUMN_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 设置标题行高度
        ws.row_dimensions[1].height = 28

        # === 写入数据行 ===
        red_font = Font(name="Microsoft YaHei", color="E74C3C", size=10)
        green_font = Font(name="Microsoft YaHei", color="27AE60", size=10)
        default_font = Font(name="Microsoft YaHei", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        # 交替行背景色
        even_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

        for row_idx, fund in enumerate(funds, start=2):
            daily_change_pct = fund.get("daily_change_pct")

            # 决定字体颜色
            if daily_change_pct is not None:
                try:
                    pct_val = float(daily_change_pct)
                    if pct_val > 0:
                        row_font = red_font
                    elif pct_val < 0:
                        row_font = green_font
                    else:
                        row_font = default_font
                except (ValueError, TypeError):
                    row_font = default_font
            else:
                row_font = default_font

            for col_idx, (key, _) in enumerate(COLUMN_HEADERS, start=1):
                value = fund.get(key, "")
                if value is None:
                    value = "--"

                # 数值列格式化
                if key in ("nav", "acc_nav", "daily_change") and value != "--":
                    try:
                        value = round(float(value), 4)
                    except (ValueError, TypeError):
                        pass
                elif key in ("daily_change_pct", "since_inception") and value != "--":
                    try:
                        value = round(float(value), 2)
                    except (ValueError, TypeError):
                        pass

                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # 涨跌着色仅应用于日增长值和日增长率列
                if key in ("daily_change", "daily_change_pct"):
                    cell.font = row_font
                else:
                    cell.font = default_font

                # 对齐方式
                if key == "name":
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                # 交替行背景
                if row_idx % 2 == 0:
                    cell.fill = even_fill

        # === 设置列宽自适应 ===
        column_widths = {
            "code": 14,
            "name": 36,
            "purchase_limit": 40,
            "index_type": 12,
            "nav": 14,
            "nav_date": 14,
            "acc_nav": 14,
            "daily_change": 14,
            "daily_change_pct": 16,
            "data_source": 14,
        }

        for col_idx, (key, _) in enumerate(COLUMN_HEADERS, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = column_widths.get(key, 14)

        # === 冻结首行 ===
        ws.freeze_panes = "A2"

        # === 设置自动筛选 ===
        if funds:
            last_col = get_column_letter(len(COLUMN_HEADERS))
            last_row = len(funds) + 1
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        # 确保目录存在
        dir_path = os.path.dirname(filepath)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)

        wb.save(filepath)
        return True

    except Exception as e:
        print(f"导出Excel失败: {e}")
        return False


def export_to_csv(funds: List[Dict], filepath: str) -> bool:
    """
    导出基金数据到 CSV 文件（UTF-8 BOM 编码，兼容 Excel）。

    Args:
        funds: 基金数据列表
        filepath: 保存文件路径

    Returns:
        True 表示成功，False 表示失败
    """
    try:
        # 确保目录存在
        dir_path = os.path.dirname(filepath)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)

            # 写入标题行
            headers = [header_text for _, header_text in COLUMN_HEADERS]
            writer.writerow(headers)

            # 写入数据行
            for fund in funds:
                row = []
                for key, _ in COLUMN_HEADERS:
                    value = fund.get(key, "")
                    if value is None:
                        value = "--"

                    # 数值列格式化
                    if key in ("nav", "acc_nav", "daily_change") and value != "--":
                        try:
                            value = f"{float(value):.4f}"
                        except (ValueError, TypeError):
                            pass
                    elif key in ("daily_change_pct", "since_inception") and value != "--":
                        try:
                            value = f"{float(value):.2f}"
                        except (ValueError, TypeError):
                            pass

                    row.append(value)

                writer.writerow(row)

        return True

    except Exception as e:
        print(f"导出CSV失败: {e}")
        return False
